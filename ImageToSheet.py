from PIL import Image
from openpyxl import Workbook

PATH_TO_IMAGE = ''
PATH_TO_FINAL_EXCEL = ''
NUMBER_OF_COLORS = 0

img = Image.open(PATH_TO_IMAGE)
pix = img.load()
workbook = Workbook()
sheet = workbook.active
w,h=img.size 
colorCounter = {str(x):0 for x in range(NUMBER_OF_COLORS)}

for i in range(w):
    for j in range(h):
        colorCounter[str(pix[i,j])] += 1
        sheet.cell(j+1,i+1,value=pix[i,j])
        

for i in range(len(colorCounter)):
    Value = '{}:{}'.format(str(i),str(colorCounter[str(i)]))
    sheet.cell(h+5,1+i,value=Value)




workbook.save(filename=PATH_TO_FINAL_EXCEL)